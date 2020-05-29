from selenium import webdriver
import time
import openpyxl


city_to_num ={
    "石家庄":1,
    "唐山":2,
    "秦皇岛":3,
    "邯郸":4,
    "邢台":5,
    "保定":6,
    "张家口":7,
    "承德市":8,
    "沧州":9,
    "廊坊":10,
    "衡水":11,
    "雄安":6
}

browser = webdriver.Chrome("D:\\OneDrive\\Programe Projects\\Python\\browserDriver\\chromedriver.exe")
browser.get("*****")

#登录
username = browser.find_element_by_xpath("//*[@id='app']/div/section/form/div[2]/div/div[1]/input")
username.send_keys("****")
password = browser.find_element_by_xpath("//*[@id='app']/div/section/form/div[3]/div/div/input")
password.send_keys("*****")
#手动输入验证码
time.sleep(10)


def createNew(phone,pswd,citynum,username,mail,msg):
    create_btn = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[1]/div/div[6]/button")
    create_btn.click()
    time.sleep(2)
    #用户名
    account = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[1]/div/div/input")
    account.send_keys(phone)

    #密码
    password = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[2]/div/div/input")
    password.send_keys(pswd)

    re_password = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[3]/div/div/input")
    re_password.send_keys(pswd)

    #获取页面的3个下拉框
    drop_downs = browser.find_elements_by_css_selector("ul[class='el-scrollbar__view el-select-dropdown__list']")
    #账号权限
    role_drop_down = browser.find_element_by_css_selector("ul[class='el-scrollbar__view el-select-dropdown__list']")
    # js1 = "document.querySelector('body > div:nth-child(7)').style.display='block';"
    # browser.execute_script(js1)
    browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[4]/div/div/div/input").click()
    time.sleep(2)
    role_drop_down.find_elements_by_tag_name("li")[4].click()

    #电话
    phonenum = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[5]/div/div[1]/input")
    phonenum.send_keys(phone)

    #省份
    province = drop_downs[1]
    # js2 = "document.querySelector('body > div:nth-child(8)').style.display='block';"
    # browser.execute_script(js2)
    browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[6]/div/div[1]/div/div[1]/input").click()
    time.sleep(2)
    province.find_element_by_tag_name("li").click()

    #地市
    city = drop_downs[2]
    # js3 = "document.querySelector('body > div:nth-child(9)').style.display='block';"
    # browser.execute_script(js3)
    browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[6]/div/div[3]/div/div/input").click()
    time.sleep(2)
    city.find_elements_by_tag_name("li")[citynum].click()


    #姓名
    name = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[7]/div/div/input")
    name.send_keys(username)

    #email
    email = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[8]/div/div[1]/input")
    email.send_keys(mail)

    #备注信息
    message = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/form/div[9]/div/div/input")
    message.send_keys(msg)

    yes_btn = browser.find_element_by_xpath(
        "//*[@id='app']/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/div/div[3]/div/button[2]")
    yes_btn.click()
    re_yes_btn = browser.find_element_by_xpath("/html/body/div[5]/div/div[3]/button[2]")
    re_yes_btn.click()


#获取账号信息
print("正在读取需要创建的账号信息")
wb=openpyxl.load_workbook("account.xlsx")
ws = wb.active
cols = ws.max_column
lines = ws.max_row
for i in range(2,lines+1):
    print("正在为 " + ws.cell(i,1).value + " 用户：" + ws.cell(i,2).value + " 创建账号。。。。。。")
    createNew(phone=ws.cell(i,4).value,
                  username=ws.cell(i,2).value,
                  mail=ws.cell(i,5).value,
                  pswd=ws.cell(i,6).value,
                  citynum= city_to_num[ws.cell(i,1).value]-1,
                  msg=ws.cell(i,3).value)
    #print(ws.cell(row=i,column=j).value)
    time.sleep(1)



